import os, re, requests, secrets, string

from django.http import JsonResponse

from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from django.conf import settings

from django.utils import timezone
from django.views.generic.base import TemplateView
from django.core.paginator import Paginator
from django.shortcuts import redirect, render
from django.urls import reverse
from django.db.models import F
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import user_passes_test
from django.utils.decorators import method_decorator
from django.contrib.auth.forms import PasswordResetForm
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_bytes, force_str
from django.db.models.query_utils import Q
from django.http import HttpResponse
from django.core.mail import send_mail, BadHeaderError
from django.template.loader import render_to_string
from django.contrib.auth.views import LoginView

from .tools import gToken



def MakeInvoiceGeneric(longitud):
    strInvoice = string.ascii_letters + string.digits
    altInvoice = ''.join(secrets.choice(strInvoice) for _ in range(longitud))
    return altInvoice

from .models import Usuario, Tickets, InvestRequests, Settings, Associate, Schedule, Messages, News

def IsStaff(user):
    return user.is_authenticated and user.is_staff

def IsDriver(user):
    return user.is_authenticated and user.is_dirver


class TestView(TemplateView):
    template_name='000.html'


class HomeView(TemplateView):
    template_name='home/home.html'

    def get(self, request, *args, **kwargs):



        try:
            ListNews = News.objects.all().order_by("-id")[:2]
        except:
            ListNews = None

        context = self.get_context_data(**kwargs)
        context={
            'ListNews':ListNews,

        }

        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):

        if 'formcontact' in request.POST:
            full_name = request.POST['fullname']
            email = request.POST['email']
            messages = request.POST['messages']

            NewMessages = Messages.objects.create(
                full_name = full_name,
                email = email,
                messages = messages
            )

            NewMessages.save()
            return redirect(reverse('Home'))

        return redirect(reverse('Home'))

class SingupView(UserPassesTestMixin, TemplateView):
    template_name='registration/singup.html'

    def test_func(self):
        """ UserAutenticate Cant Create NewUser, When User is Antenticate test_func is False
        """
        return not self.request.user.is_authenticated

    def handle_no_permission(self):
        """When test_fun(self) is False, Redirect User Autenticate to Home
        """
        return redirect(reverse('Home'))

    def post(self, request, *args, **kwargs):
        """Tasker request Form to Create a NewUser, Verifi User/Email in DB 
            If Success form Send Email Verification to User
        """ 
        iUsername = request.POST['username']
        iPass = request.POST['password']
        iFullName = request.POST['name']
        iEmail = request.POST['email']

        LastID = Usuario.objects.latest('id').id + 1

        if not re.match(r'^[a-zA-Z0-9]+$', iUsername):
            messages.error(request, '¡Registro Incompleto!', extra_tags="title")
            messages.error(request, f'El Nombre de Usuario no es Valido', extra_tags="info") 
            return redirect(reverse('Singup'))

        if Usuario.objects.filter(username=iUsername):
            messages.error(request, '¡Registro Incompleto!', extra_tags="title")
            messages.error(request, f'El Nombre de Usuario no esta Disponible', extra_tags="info") 
            return redirect(reverse('Singup'))

        if Usuario.objects.filter(email=iEmail, is_active=True):
            messages.error(request, '¡Registro Incompleto!', extra_tags="title")
            messages.error(request, f'El Email no esta Disponible', extra_tags="info") 
            return redirect(reverse('Singup'))
        
        
        request.session['django_messages'] = []

        try:
            nUser = Usuario.objects.create(
                username = iUsername,
                full_name = iFullName,
                email = iEmail,
                codigo = LastID
            )
            
            nUser.set_password(iPass)
            nUser.save()

            #login(request, nUser) -->Email ServiceAutenticate
                        


            messages.success(request, '¡Registro Exitoso!', extra_tags="title")
            messages.success(request, f'Activaremos tu cuenta en las siguientes 48 Horas', extra_tags="info")
            
        except Exception as e:
            messages.error(request, '¡Registro Incompleto!', extra_tags="title")
            messages.error(request, f'Ha ocurrido un error durante el registro', extra_tags="info")    

        return redirect(reverse('Singup'))

class UserLoginView(UserPassesTestMixin, LoginView):
    template_name='registration/login.html'

    def test_func(self):
        """ UserAutenticate Cant Create NewUser, When User is Antenticate test_func is False
        """
        return not self.request.user.is_authenticated

    def handle_no_permission(self):
        """When test_fun(self) is False, Redirect User Autenticate to Home
        """
        return redirect(reverse('Home'))


    def form_invalid(self, form):
        messages.error(self.request, 'Usuario/Contraseña Incorrectos', extra_tags="title")
        messages.error(self.request, 'Intentelo Nuevamente', extra_tags="info")
        return super().form_invalid(form)


class InvestmentView(LoginRequiredMixin, TemplateView):
    template_name='home/investment.html'

class InvestPremiumView(LoginRequiredMixin, TemplateView):
    template_name='home/invest_premium.html'
    
    @method_decorator(user_passes_test(IsStaff))
    def dispatch(self, request, *args, **kwargs):
        """ Only UserStaff Can Acces to this View
        """
        return super().dispatch(request, *args, **kwargs)

class InfoFormView(LoginRequiredMixin, TemplateView):
    template_name='home/info_ticket.html'


class AdminServicesHistory(LoginRequiredMixin, TemplateView):
    template_name='driver/history.html'

    @method_decorator(user_passes_test(IsDriver))
    def dispatch(self, request, *args, **kwargs):
        """ Only Drivers Can Acces to this View
        """
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):

        ITEMS = 5
        MAXPAGES = 5

        iSchedule = Schedule.objects.filter(~Q(status="Pendiente") & Q(driver=request.user.codigo)).order_by('-id')[:ITEMS*MAXPAGES]
        ListSchedule = Paginator(iSchedule,ITEMS).get_page(request.GET.get('page')) if iSchedule else []
        
        ListFix = ITEMS - len(iSchedule)%ITEMS

        if ListFix == ITEMS and len(iSchedule) != 0:
            ListFix = 0

        context = self.get_context_data(**kwargs)
        context={
            'ListSchedule':ListSchedule,
            'ListFix':range(0,ListFix),
        }
        return self.render_to_response(context)


class AdminServicesUser(LoginRequiredMixin, TemplateView):
    template_name='driver/admin.html'

    @method_decorator(user_passes_test(IsDriver))
    def dispatch(self, request, *args, **kwargs):
        """ Only Drivers Can Acces to this View
        """
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):

        if 'cancel' in request.POST:
            iCode = request.POST['iCode']
            CUser = Usuario.objects.get(codigo=iCode)
            TSchedule = Schedule.objects.filter(Q(status="Pendiente") & Q(username=CUser)).order_by('-id').first()

            messages.success(request, '¡Servicio Cancelado!', extra_tags="title")
            messages.success(request, f'El servicio ha sido marcado como cancelado', extra_tags="info")
            iUser = Usuario.objects.filter(id=request.user.id)
            iUser.update(is_driving=False)

            TSchedule.status = "Cancelado"
            TSchedule.paid = 0
            TSchedule.save()

            return redirect(reverse('svAdminUser'))

        if 'input' in request.POST:
            iCode = request.POST['iCode']

            try:
                InfoUser = Usuario.objects.get(codigo=iCode)
                UserShedule = Schedule.objects.filter(Q(status="Pendiente") & Q(username=InfoUser)).order_by('-id').first()
                if UserShedule:
                    TSchedule = UserShedule
                else:
                    TSchedule = Schedule.objects.create(
                        username = InfoUser,
                        driver = request.user.codigo,
                        date = timezone.now(),
                        status = "Pendiente",
                        )

                    TSchedule.save() 


                context = self.get_context_data(InfoUser=Usuario.objects.get(codigo=iCode),TSchedule=TSchedule)
                return self.render_to_response(context)

            except ObjectDoesNotExist:
                messages.error(request, '¡Usuario Inexistente!', extra_tags="title")
                messages.error(request, f'El Usuario ingresado no corresponde a un Afiliado', extra_tags="info")
                return redirect(reverse('svAdminUser'))

        if 'success' in request.POST:
            iCode = request.POST['iCode']
            iUser = Usuario.objects.get(id=request.user.id)
            iUser.is_driving = True
            iUser.save()

            InfoUser = Usuario.objects.get(codigo=iCode)
            TSchedule = Schedule.objects.filter(Q(status="Pendiente") & Q(username=InfoUser)).order_by('-id').first()

            context = self.get_context_data(InfoUser=InfoUser,TSchedule=TSchedule,IsDriving=True)
            return self.render_to_response(context)


        if 'fonds' in request.POST:
            iCode = request.POST['iCode']
            iValue = int(request.POST['iValue'])
            CUser = Usuario.objects.get(codigo=iCode)

            ifrom = request.POST['ifrom']
            ito = request.POST['ito']
            idistance = int(request.POST['idistance'])

            TSchedule = Schedule.objects.filter(Q(status="Pendiente") & Q(username=CUser)).order_by('-id').first()

            try:

                if CUser.available > iValue:
                    CUser.available -= iValue
                    CUser.save()
                    messages.success(request, '¡Servicio Completado!', extra_tags="title")
                    messages.success(request, f'El pago se ha completado satisfactoriamente', extra_tags="info")
                    iUser = Usuario.objects.filter(id=request.user.id)
                    iUser.update(is_driving=False)

                    TSchedule.status = "Completado"
                    TSchedule.paid = iValue
                    TSchedule.addres_from = ifrom
                    TSchedule.addres_to = ito
                    TSchedule.distance = idistance
                    TSchedule.save()

                    return redirect(reverse('svAdminUser'))

                else:
                    messages.error(request, '¡Fondos Insuficientes!', extra_tags="title")
                    messages.error(request, f'Seleccione otro metodo de pago', extra_tags="info")
                    return redirect(reverse('svAdminUser'))

            except Exception as e:
                with open(os.path.join(settings.BASE_DIR, 'logs/logdriver.txt'), 'a') as f:
                    f.write("{} QueryError Interest: {}\n".format(str(CUser.username), str(e)))


        if 'cash' in request.POST:

            iCode = request.POST['iCode']
            CUser = Usuario.objects.get(codigo=iCode)
            iValue = int(request.POST['iValue'])
            ifrom = request.POST['ifrom']
            ito = request.POST['ito']
            idistance = int(request.POST['idistance'])

            TSchedule = Schedule.objects.filter(Q(status="Pendiente") & Q(username=CUser)).order_by('-id').first()

            try:
                messages.success(request, '¡Servicio Completado!', extra_tags="title")
                messages.success(request, f'El pago se ha confirmado como efectivo', extra_tags="info")
                iUser = Usuario.objects.filter(id=request.user.id)
                iUser.update(is_driving=False)

                TSchedule.status = "Completado"
                TSchedule.paid = iValue
                TSchedule.addres_from = ifrom
                TSchedule.addres_to = ito
                TSchedule.distance = idistance
                TSchedule.save()

                return redirect(reverse('svAdminUser'))

            except Exception as e:
                with open(os.path.join(settings.BASE_DIR, 'logs/logdriver.txt'), 'a') as f:
                    f.write("{} QueryError Interest: {}\n".format(str(CUser.username), str(e)))

        return redirect(reverse('svAdminUser'))

    def get(self, request, *args, **kwargs):

        Sumbmit = request.GET.get('submit')
        context = self.get_context_data(**kwargs)


        if Sumbmit == 'clear':
            iUser = Usuario.objects.filter(id=request.user.id)
            iUser.update(is_driving=False)
            context = {
                'ListSchedule': None,
                'TSchedule': None,
                'InfoUser': None,
            }
            return self.render_to_response(context)

        return self.render_to_response(context)

class AdminServices(LoginRequiredMixin, TemplateView):
    template_name='driver/home.html'

    @method_decorator(user_passes_test(IsDriver))
    def dispatch(self, request, *args, **kwargs):
        """ Only Drivers Can Acces to this View
        """
        return super().dispatch(request, *args, **kwargs)


    def get(self, request, *args, **kwargs):

        ITEMS = 5
        MAXPAGES = 5

        iSchedule = Schedule.objects.filter(Q(status="Pendiente") & Q(driver=request.user.codigo)).order_by('id')[:ITEMS*MAXPAGES]
        cSchedule = Paginator(iSchedule,ITEMS).get_page(request.GET.get('page')) if iSchedule else []
        
        ListFix = ITEMS - len(iSchedule)%ITEMS

        if ListFix == ITEMS and len(iSchedule) != 0:
            ListFix = 0

        context = self.get_context_data(**kwargs)
        context={
            'cSchedule':cSchedule,
            'ListFix':range(0,ListFix),
        }
        return self.render_to_response(context)

class AdminServicesAdd(LoginRequiredMixin, TemplateView):
    template_name='driver/add.html'

    def post(self, request, *args, **kwargs):

        if 'input' in request.POST:
            iCode = request.POST['iCode']

            try:
                context = self.get_context_data(InfoUser=Usuario.objects.get(codigo=iCode))
                return self.render_to_response(context)

            except ObjectDoesNotExist:
                messages.error(request, '¡Usuario Inexistente!', extra_tags="title")
                messages.error(request, f'El Usuario ingresado no corresponde a un Afiliado', extra_tags="info")
                return redirect(reverse('svAdminUserAdd'))
            
        if 'add' in request.POST: 
            iCode = request.POST['iCode']
            iFrom = request.POST['ifrom']
            iTo = request.POST['ito']
            iDate = request.POST['idate']

            try:
                InfoUser = Usuario.objects.get(codigo=iCode)
                TSchedule = Schedule.objects.create(
                    username = InfoUser,
                    driver = request.user.codigo,
                    date = timezone.make_aware(datetime.strptime(iDate, "%Y-%m-%dT%H:%M"), timezone.get_current_timezone()),
                    addres_from = iFrom,
                    addres_to = iTo,
                    status = "Pendiente",
                    )

                TSchedule.save() 

                messages.success(request, 'Solicitud Registrada', extra_tags="title")
                messages.success(request, f'Agendamiento Exitoso, El Registro se ha completado', extra_tags="info")
                return redirect(reverse('svAdminAdd'))

            except ObjectDoesNotExist:
                messages.error(request, '¡Usuario Inexistente!', extra_tags="title")
                messages.error(request, f'El Usuario ingresado no corresponde a un Afiliado', extra_tags="info")
                return redirect(reverse('svAdminAdd'))
            
        return redirect(reverse('svAdminAdd'))

class InfoView(TemplateView):
    template_name='home/info.html'
    
    def post(self, request, *args, **kwargs):

        base_currency = 'USD'
        target_currency = 'COP'

        APILAYER = settings.APILAYER_KEY
        url = f'https://api.exchangeratesapi.io/v1/convert?access_key={APILAYER}&from={base_currency}&to={target_currency}&amount={1}'

        try:
            response = requests.get(url)
            data = response.json()
            usd_to_cop = data['result'] if 'result' in data else Settings.objects.get(Online=True).exchange
            
        except requests.exceptions.RequestException as e:
            return JsonResponse({'Error': 'Conversion Fallida'}, status=500)

        InfoUser = Usuario.objects.get(id=request.user.id)
        iUsername = InfoUser.username

        if InfoUser.is_operating:
            messages.error(request, 'ERROR', extra_tags="title")
            messages.error(request, 'Actualmente ya cuenta con una Inversión Activa', extra_tags="info")
            return redirect(reverse('Info')) 

        rInstanceUser = Usuario.objects.get(username=iUsername)
        
        latest_invest_request = InvestRequests.objects.last()
        iCode = latest_invest_request.pk if latest_invest_request is not None else 0


        iName = InfoUser.full_name
        iEmail = InfoUser.email
        iCountry = request.POST['country']
        iPhone = request.POST['phone']
        iCurrency = str(request.POST['currency'])

        iAmount = int(request.POST['amount'])
        iAmount_to_cop = iAmount if iCurrency == "COP" else iAmount*usd_to_cop

        iBank = str(request.POST['bank'])
        iBankAccount = request.POST['bank_account']
        
        Interest = 3
        iDateExpire = timezone.now() + timedelta(days=365)

        iDateString = iDateExpire.isoformat()
        iDateObject = datetime.fromisoformat(iDateString)


        try:
            data = {
                'codigo': iCode,
                'full_name': iName,
                'email': iEmail,
                'country': iCountry,
                'phone': iPhone,
                'ammount': iAmount_to_cop,
                'bank': iBank,
                'bank_account': iBankAccount,
                'staff': "vaor",
                'staff_cod': 2311619210000,
                'date_joined': timezone.now(),
                'interest': Interest,
                'date_expire': iDateObject,
                'rState': "Pendiente"
            }

            obj, created = InvestRequests.objects.update_or_create(username=rInstanceUser,defaults=data)

            subject = "Solicitud - Información Inversión"        
            email_template_name = "home/email/info_email.txt"

            c = {
                'iUsername': iUsername,
                'iName':iName,
                'iAmount':iAmount_to_cop,
                'iBank':iBank,   
                'site_name': 'VRT-Fund',
                'protocol': 'https',# http
                'domain':'vrtfund.com',# 127.0.0.1:8000
            }
            email = render_to_string(email_template_name, c)

            try:
                send_mail(subject, email, 'noreply@vrtfund.com' , [email], fail_silently=False)
            except Exception as e:
                with open(os.path.join(settings.BASE_DIR, 'logs/email.log.txt'), 'a') as f:
                    eDate = timezone.now().strftime("%Y-%m-%d %H:%M")
                    f.write("EmailError InfoEmail--> {} Error: {}\n".format(eDate, str(e)))
            
            APIAmount = str("{:.2f}".format(iAmount)) if iCurrency == "USD" else str("{:.2f}".format(iAmount/usd_to_cop))

            body = {
                "product": {
                    "description": "VRTFund",
                    "name": "Investment"
                },
                "invoice": {
                    "amount": APIAmount,
                    "currencyFrom": "USD"
                },
                "settlement": {
                    "currency": "USD"
                },
                "notifyEmail": "noreply@vrtfund.com",
                "notifyUrl": "https://vrtfund.com/",
                "returnUrl": "https://vrtfund.com/",
                "reference": "anything"
            }

            CONFIRMO = settings.CONFIRMO_KEY
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {CONFIRMO}'

            }

            try:
                response = requests.post('https://confirmo.net/api/v3/invoices', json=body, headers=headers)
                Invoice = response.json().get('id') if response.status_code == 201 else MakeInvoiceGeneric(12)
                InvRequest = InvestRequests.objects.get(username=self.request.user)
                InvRequest.invoice = Invoice
                InvRequest.save()

                return redirect(reverse('Payments'))
                
            except Exception as e:
                messages.error(request, 'ERROR', extra_tags="title")
                messages.error(request, f'La solicitud no se ha podido procesar {data}', extra_tags="info")
                return redirect(reverse('InfoForm')) 
            
        except Exception as e:
            with open(os.path.join(settings.BASE_DIR, 'logs/email.log'), 'a') as f:
                eDate = timezone.now().strftime("%Y-%m-%d %H:%M")
                f.write("QueryError InfoForm--> {} Error: {}\n".format(eDate, str(e)))
            
            messages.error(request, 'ERROR', extra_tags="title")
            messages.error(request, f'La solicitud no se ha podido procesar {data}', extra_tags="info")
            return redirect(reverse('InfoForm')) 



class PaymentsView(LoginRequiredMixin, TemplateView):
    template_name='home/payment.html'

    def get(self, request, *args, **kwargs):
        
        InvRequest = InvestRequests.objects.get(username=self.request.user)
        
        IntsDayli = InvRequest.interest/(100*30)
        TimeDelta = (InvRequest.date_expire - InvRequest.date_joined).days

        Monthly = int(InvRequest.ammount*IntsDayli*30)
        MaxAmount = int(InvRequest.ammount * TimeDelta * IntsDayli)

        context = self.get_context_data(**kwargs)
        context={
            'Monthly':Monthly,
            'MaxAmount':MaxAmount,
            'InvRequest':InvRequest,
        }

        return self.render_to_response(context)


class PaymentsBanks(LoginRequiredMixin, TemplateView):
    template_name='home/payment-info.html'

    def get(self, request, *args, **kwargs):
        
        InvRequest = InvestRequests.objects.get(username=self.request.user)
        
        IntsDayli = InvRequest.interest/(100*30)
        TimeDelta = (InvRequest.date_expire - InvRequest.date_joined).days

        Monthly = int(InvRequest.ammount*IntsDayli*30)
        MaxAmount = int(InvRequest.ammount * TimeDelta * IntsDayli)

        context = self.get_context_data(**kwargs)
        context={
            'Monthly':Monthly,
            'MaxAmount':MaxAmount,
            'InvRequest':InvRequest,
        }

        return self.render_to_response(context)

#LoginRequiredMixin
class ContentView(TemplateView):
    template_name='home/content.html'
    
#LoginRequiredMixin
class BenefitView(TemplateView):
    template_name='home/benefit.html'

class ServicesView(TemplateView):
    template_name='home/services.html'

    def get(self, request, *args, **kwargs):
        
        ListGift = Associate.objects.filter(IsActive=True).order_by("id")

        context = self.get_context_data(**kwargs)
        context={
            'ListGift':ListGift,
        }

        return self.render_to_response(context)
    

class InterfaceView(LoginRequiredMixin, TemplateView):
    template_name='interface/interface.html'

class LegalView(TemplateView):
    template_name='home/legal.html'
        

class TicketFormView(LoginRequiredMixin, TemplateView):
    template_name='interface/tickets.html'
    

class HistoryListView(LoginRequiredMixin, TemplateView):
      
    template_name='interface/history.html'

    def days_until_next_month(self):
        Today = timezone.now()
        NextMonth = Today.replace(day=28) + timedelta(days=4)
        TimeDelta = NextMonth.replace(day=1) - Today
        return TimeDelta.days

    def get(self, request, *args, **kwargs):
        
        ITEMS = 5
        MAXPAGES = 5
        
        OTickets = Tickets.objects.filter(username=request.user.id).order_by("-date")[:ITEMS*MAXPAGES]
        ListTicketsPages = Paginator(OTickets,ITEMS).get_page(request.GET.get('page'))
        
        TicketsFix = ITEMS - len(OTickets)%ITEMS
        
        if TicketsFix == ITEMS and len(OTickets) != 0:
            TicketsFix = 0
        
        context = self.get_context_data(**kwargs)
        context={
            'ListTicketsPages':ListTicketsPages,
            'FixTicketsPage':range(0,TicketsFix)
        }

        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):

        InfoUser = Usuario.objects.get(id=request.user.id)
        AviableTickets = InfoUser.available_tickets

        try:
            Setting = Settings.objects.get(Online=True)  
        except:
            Setting = None
            
        Fee = Setting.sFee if Setting else 0

        rAmmount = int(request.POST['ammount'])
        rAmmountFrom = request.POST['ammount_from']
        rBank = request.POST['bank']
        rBankAccount = request.POST['bank_account']
        rComment = request.POST["message"]
        rState = "Pendiente"
        
        if AviableTickets < 1:
            messages.error(request, 'ERROR', extra_tags="title")
            messages.error(request, '¡Ha excedido el número de retiros Mensuales!', extra_tags="info")
            return redirect(reverse('History'))

        if rAmmountFrom == "f1":
            rAmmountFrom = "Intereses"
            
            if rAmmount > InfoUser.available:
                messages.error(request, 'ERROR', extra_tags="title")
                messages.error(request, 'La solicitud no se ha podido procesar', extra_tags="info")
                return redirect(reverse('History'))
            
            Usuario.objects.filter(id=InfoUser.id).update(
                available=F('available')-rAmmount,
                paid=F('paid')+rAmmount
                )

        if rAmmountFrom == "f2":
            rAmmountFrom = "Comisiones"
            
            if rAmmount > InfoUser.ref_available:
                messages.error(request, 'ERROR', extra_tags="title")
                messages.error(request, 'La solicitud no se ha podido procesar', extra_tags="info")
                return redirect(reverse('History'))
            
            Usuario.objects.filter(id=InfoUser.id).update(
                ref_available=F('ref_available')-rAmmount,
                ref_paid=F('ref_paid')+rAmmount
                )

        rAmmountFee = rAmmount - Fee
        
        Tickets.objects.create(
            username = InfoUser,
            tAmmount = rAmmountFee,
            tAmmountFrom = rAmmountFrom,
            tBank= rBank,
            tBankAccount = rBankAccount,
            CommentText = rComment,
            rState=rState
            ) 

        TimeDelta = self.days_until_next_month()

        subject = "Notificación - Solicitud de Retiro"        
        email_template_name = "interface/email/tickets_email_notify.html"

        c = {
        'username': InfoUser.username,
        'tAmmount':rAmmount,
        'tAmmountFrom':rAmmountFrom,
        'tBank':rBank,
        'tBankAccount':rBankAccount,   
        'TimeDelta':TimeDelta,    
        'site_name': 'VRT-Fund',
        'protocol': 'https',# http
        'domain':'vrtfund.com',# 127.0.0.1:8000
        }
        email = render_to_string(email_template_name, c)

        try:
            send_mail(subject, message=None, from_email='noreply@vrtfund.com', recipient_list=[InfoUser.email], fail_silently=False, html_message=email)
        except Exception as e:
            with open(os.path.join(settings.BASE_DIR, 'logs/email.log'), 'a') as f:
                eDate = timezone.now().strftime("%Y-%m-%d %H:%M")
                f.write("EmailTicket Notification--> {} Error: {}\n".format(eDate, str(e)))


        FileName = os.path.join(settings.BASE_DIR, 'logs/users/') + InfoUser.username + '.xlsx'

        try:
            if not os.path.exists(FileName):
                WB = Workbook()
                WS = WB.active
                WS.append(["Tipo","Fecha","$Interes","$Comiciones","AcInteres","AcComisiones","$Ticket","Origen","Total","VRTs Acumulados","VRTs Usados","VRTs Totales"])
            else:
                WB = load_workbook(FileName)
                WS = WB.active

            NowToday = timezone.now().strftime("%Y-%m-%d %H:%M")

            FileData = [0, NowToday, "", "", "", "", rAmmount, rAmmountFrom,"","","",""]

            WS.append(FileData)
            WB.save(FileName)
            
        except Exception as e:
            with open(os.path.join(settings.BASE_DIR, 'logs/workbook.txt'), 'a') as f:
                f.write("HistoryList WorkbookError: {}\n".format(str(e)))


        Usuario.objects.filter(id=InfoUser.id).update(available_tickets=F('available_tickets')-1)

        try:
            Setting = Settings.objects.get(Online=True) 
            Setting.sFeeAmmount += Fee
            Setting.save()
             
        except Settings.DoesNotExist:
            Setting = None

        
        messages.success(request, 'Solicitud Registrada', extra_tags="title")
        messages.success(request, f'¡EL tiempo de espera aproximado será de {TimeDelta} días Hábiles!', extra_tags="info")
        return redirect(reverse('History'))


def PasswordResetRequestView(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            associated_users = Usuario.objects.filter(Q(email=data))
            if associated_users.exists():
                for user in associated_users:
                    subject = "Solicitud - Restablecer Contraseña"
                    email_template_name = "password/email/password_reset_email.txt"
                    c = {
                    'username': user.username,
                    "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                    "user": user,
                    'token': default_token_generator.make_token(user),
                    'site_name': 'VRT-Fund',
                    'protocol': 'https',# http
                    'domain':'vrtfund.com',# 127.0.0.1:8000
					}
                    email = render_to_string(email_template_name, c)
                    try:
                        send_mail(subject, email, 'noreply@vrtfund.com' , [user.email], fail_silently=False)
                    except BadHeaderError:
                        return HttpResponse('Invalid Header Found')
                    return redirect ("/accounts/password_reset/done/")
 
    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="password/password_reset.html", context={"password_reset_form":password_reset_form})



def EmailConfirmView(request, uidb64, token):
   
        try:
            uid = force_str(urlsafe_base64_decode(uidb64))
            nUser = Usuario.objects.get(pk=uid)

            if nUser and gToken.check_token(nUser, token):
                nUser.is_active = True
                nUser.save()

                return render(request, 'registration/email_confirm.html', {"user": nUser})
            
        except Exception as e:
            nUser = None
            with open(os.path.join(settings.BASE_DIR, 'logs/email.log'), 'a') as f:
                eDate = timezone.now().strftime("%Y-%m-%d %H:%M")
                f.write("EmailConfirm--> {} Error: {}\n".format(eDate, str(e)))

        return render(request, 'registration/email_confirm-failed.html', {"user": nUser})

def ComingSoonView(request):
    return render(request, '404.html')


class GiftView(LoginRequiredMixin, TemplateView):
    template_name='gift/gift.html'

    def get(self, request, *args, **kwargs):
        
        ListGift = Associate.objects.all().order_by("id")

        context = self.get_context_data(**kwargs)
        context={
            'ListGift':ListGift,
        }

        return self.render_to_response(context)

class GiftTicketView(LoginRequiredMixin, TemplateView):
    template_name='gift/giftticket.html'

    def get(self, request, *args, **kwargs):
        
        ListGift = Associate.objects.filter(IsActive=True).order_by("id")

        context = self.get_context_data(**kwargs)
        context={
            'ListGift':ListGift,
        }

        return self.render_to_response(context)


