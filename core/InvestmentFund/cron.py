import os

from openpyxl import Workbook
from openpyxl import load_workbook

from datetime import date
from django.conf import settings

from django_cron import CronJobBase, Schedule
from django.db.models import F
from django.utils import timezone

from .models import Usuario


class AddFundsToUser(CronJobBase):
    RUN_EVERY_MINS = 1

    schedule = Schedule(run_every_mins=RUN_EVERY_MINS)
    code = 'InvestmentFund.AddFundsToUser'
    
    #crontab -e
    #* * * * * /home/savelasquezo/savelasquezo/VaorTrading/venv/bin/python /home/savelasquezo/savelasquezo/VaorTrading/core/manage.py runcrons
    def do(self):
        
        NowToday = timezone.now().strftime("%Y-%m-%d %H:%M")
        
        with open(os.path.join(settings.BASE_DIR, 'logs/logcron.txt'), 'a') as f:
            f.write("ServiceCron Active {}\n".format(NowToday))

        InfoUser = Usuario.objects.all()
        #Usuario.objects.filter(date_expire__lte=timezone.now()).update(is_operating=False)

        for nUser in InfoUser:

            if nUser.is_operating:

                CUser = Usuario.objects.filter(id=nUser.id)

                if date.today().day == 1:
                    CUser.update(available_tickets=3)

                cUsername = nUser.username
                cAmmount = int(nUser.ammount)
                cTotalInterest = int(nUser.total_interest)
                
                cPaid = int(nUser.paid)
                cPaidRef = int(nUser.ref_paid)

                cInterest = float((nUser.interest)/(100*30))
                cInterestRef = float((nUser.ref_interest)/(100*30))

                cValue = int(cAmmount*(cInterest))
                cValueRef = int(cAmmount*(cInterestRef))
                
                cAvailable = int(cTotalInterest-cPaid + cValue)

                try:
                    CUser.update(
                        available=cAvailable,
                        total_interest=F('total_interest') + cValue)
                    
                except Exception as e:
                    with open(os.path.join(settings.BASE_DIR, 'logs/logcron.txt'), 'a') as f:
                        f.write("{} QueryError Interest: {}\n".format(str(cUsername), str(e)))


                try:
                    if nUser.ref_id:
                        cValueRef = int(cAmmount*(cInterestRef))
                    CUser.update(ref_total=F('ref_total') + cValueRef)
                    
                except Exception as e:
                    with open(os.path.join(settings.BASE_DIR, 'logs/logcron.txt'), 'a') as f:
                        f.write("QueryError Referido: {}\n".format(str(e)))
                
                UserRef = Usuario.objects.filter(ref_id= nUser.codigo)
                mAviableUserRef = 0
                
                for mUser in UserRef:
                    mAviableUserRef += int(mUser.ref_total)

                mAviableUserTotal = mAviableUserRef - cPaidRef
                cTodayRef =  mAviableUserTotal - nUser.ref_available
                
                try:
                    CUser.update(ref_available=mAviableUserTotal,total_ref=mAviableUserRef)
                    CUser.update(total=F('total_ref') + F('total_interest'))
                    
                except Exception as e:
                    with open(os.path.join(settings.BASE_DIR, 'logs/logcron.txt'), 'a') as f:
                        f.write("QueryError Asociados: {}\n".format(str(e)))   
                        
                FileName = '/home/savelasquezo/apps/vrt/core/logs/users/'+ nUser.username + '.xlsx'
                try:
                    if not os.path.exists(FileName):
                        WB = Workbook()
                        WS = WB.active
                        WS.append(["Tipo","Fecha","$Interes","$Comiciones","AcInteres","AcComisiones","$Ticket","Origen","Total"])
                    else:
                        WB = load_workbook(FileName)
                        WS = WB.active
                    
                    cTotal = nUser.total
                    cAviableRef = nUser.ref_available

                    FileData = [1, NowToday, cValue, cTodayRef, cAvailable, cAviableRef, "", "", cTotal]

                    WS.append(FileData)
                    WB.save(FileName)
                    
                except Exception as e:
                    with open(os.path.join(settings.BASE_DIR, 'logs/workbook.txt'), 'a') as f:
                        f.write("CronJob WorkbookError: {}\n".format(str(e)))

cronjobs = [
    AddFundsToUser,
]